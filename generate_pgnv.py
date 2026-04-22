import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: Cử nhân
# ============================================================
ws = wb.active
ws.title = "Cử nhân"

# Column widths (approximate matching)
col_widths = {
    'A': 4, 'B': 12, 'C': 18, 'D': 10, 'E': 8, 'F': 8,
    'G': 8, 'H': 10, 'I': 14, 'J': 8, 'K': 4
}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# Row heights
ws.row_dimensions[1].height = 18
ws.row_dimensions[5].height = 24

# Fonts
font_normal = Font(name='Times New Roman', size=11)
font_bold = Font(name='Times New Roman', size=11, bold=True)
font_title = Font(name='Times New Roman', size=13, bold=True)
font_header = Font(name='Times New Roman', size=12, bold=True)
font_small = Font(name='Times New Roman', size=10)
font_italic = Font(name='Times New Roman', size=11, italic=True)
font_small_italic = Font(name='Times New Roman', size=10, italic=True)

# Alignments
align_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_right = Alignment(horizontal='right', vertical='top', wrap_text=True)

# Set default font for all cells
for row in range(1, 120):
    for col in range(1, 12):
        ws.cell(row=row, column=col).font = font_normal
        ws.cell(row=row, column=col).alignment = align_left

# ---- Row 1: Mẫu ĐATN 02 ----
ws.merge_cells('J1:K1')
ws['J1'] = "Mẫu ĐATN 02"
ws['J1'].font = font_small_italic
ws['J1'].alignment = Alignment(horizontal='right')

# ---- Row 2-3: Header ----
ws.merge_cells('A2:K2')
ws['A2'] = "ĐẠI HỌC BÁCH KHOA HÀ NỘI"
ws['A2'].font = font_bold
ws['A2'].alignment = align_center

ws.merge_cells('A3:K3')
ws['A3'] = "TRƯỜNG CÔNG NGHỆ THÔNG TIN VÀ TRUYỀN THÔNG"
ws['A3'].font = font_bold
ws['A3'].alignment = align_center

# ---- Row 5: Main Title ----
ws.merge_cells('A5:K5')
ws['A5'] = "PHIẾU GIAO NHIỆM VỤ ĐỒ ÁN TỐT NGHIỆP HỆ CỬ NHÂN"
ws['A5'].font = font_title
ws['A5'].alignment = align_center

# ---- Row 6: Kỳ ----
ws['E6'] = "KỲ"
ws['E6'].font = font_bold
ws['F6'] = "2025.2"
ws['F6'].font = font_bold

# ---- Row 8: Thông tin sinh viên ----
ws.merge_cells('A8:K8')
ws['A8'] = "Thông tin về sinh viên"
ws['A8'].font = font_header

# ---- Row 9: Họ tên + MSSV ----
ws.merge_cells('A9:B9')
ws['A9'] = "Họ và tên sinh viên:"
ws['A9'].font = font_bold
ws.merge_cells('C9:G9')
ws['C9'] = "Dương Phương Thảo"
ws['C9'].font = font_normal
ws['H9'] = "MSSV:"
ws['H9'].font = font_bold
ws.merge_cells('I9:K9')
ws['I9'] = "20226001"
ws['I9'].font = font_normal

# ---- Row 10: ĐT + Lớp ----
ws.merge_cells('A10:B10')
ws['A10'] = "Điện thoại liên lạc:"
ws['A10'].font = font_bold
ws.merge_cells('C10:G10')
ws['C10'] = "0913627734"
ws['C10'].font = font_normal
ws['H10'] = "Lớp:"
ws['H10'].font = font_bold
ws.merge_cells('I10:K10')
ws['I10'] = "ITE7-02-K67"
ws['I10'].font = font_normal

# ---- Row 11: Email + Mã lớp ----
ws.merge_cells('A11:B11')
ws['A11'] = "Email:"
ws['A11'].font = font_bold
ws.merge_cells('C11:G11')
ws['C11'] = "thao.dp226001@sis.hust.edu.vn"
ws['C11'].font = font_normal
ws['H11'] = "Mã lớp:"
ws['H11'].font = font_bold
ws.merge_cells('I11:K11')
ws['I11'] = ""
ws['I11'].font = font_normal

# ---- Row 12: Thông tin GVHD ----
ws.merge_cells('A12:K12')
ws['A12'] = "Thông tin giáo viên hướng dẫn"
ws['A12'].font = font_header

# ---- Row 13: GVHD ----
ws.merge_cells('A13:B13')
ws['A13'] = "Họ và tên GVHD:"
ws['A13'].font = font_bold
ws.merge_cells('C13:K13')
ws['C13'] = "Phạm Quang Dũng"
ws['C13'].font = font_normal

# ---- Row 14: Nơi thực hiện ----
ws.merge_cells('A14:B14')
ws['A14'] = "Đồ án được thực hiện tại:"
ws['A14'].font = font_bold
ws.merge_cells('C14:K14')
ws['C14'] = "Trường Công nghệ thông tin và Truyền thông"
ws['C14'].font = font_normal

# ---- Row 15: Thời gian ----
ws.merge_cells('A15:B15')
ws['A15'] = "Thời gian làm ĐATN: "
ws['A15'].font = font_bold
ws['C15'] = "Từ ngày"
ws['C15'].font = font_normal
ws['D15'] = "24/2/2026"
ws['D15'].font = font_normal
ws.merge_cells('H15:H15')
ws['H15'] = "đến ngày"
ws['H15'].font = font_normal
ws['I15'] = "15/6/2026"
ws['I15'].font = font_normal

# ---- Row 16-17: Tên đề tài ----
ws.merge_cells('A16:K16')
ws['A16'] = "1. Tên đề tài:"
ws['A16'].font = font_bold

ws.merge_cells('A17:K17')
ws['A17'] = "Phần mềm quản lý chuỗi cung ứng thông minh"
ws['A17'].font = font_normal

# ---- Row 18-22: Lĩnh vực ----
ws.merge_cells('A18:K18')
ws['A18'] = "2. Lĩnh vực đề tài:"
ws['A18'].font = font_bold

ws['A19'] = "-"
ws['B19'] = "Lựa chọn 1:"
ws['B19'].font = font_bold
ws.merge_cells('C19:K19')
ws['C19'] = "Thương mại điện tử (eCommerce) và hậu cần (Logistic)"
ws['C19'].font = font_normal

ws['A20'] = "-"
ws['B20'] = "Lựa chọn 2:"
ws['B20'].font = font_bold
ws.merge_cells('C20:K20')
ws['C20'] = "Phần mềm doanh nghiệp"
ws['C20'].font = font_normal

ws['A21'] = "-"
ws['B21'] = "Lựa chọn 3:"
ws['B21'].font = font_bold
ws.merge_cells('C21:K21')
ws['C21'] = ""

ws['A22'] = "-"
ws.merge_cells('B22:K22')
ws['B22'] = "Nếu lĩnh vực không nằm trong danh sách có sẵn, giáo viên hướng dẫn có thể đề xuất:"
ws['B22'].font = font_italic

# ---- Row 24: Mục tiêu ----
ws.merge_cells('A24:K24')
ws['A24'] = "3. Mục tiêu của ĐATN:"
ws['A24'].font = font_bold

# 3.1 Kiến thức
ws.merge_cells('A25:K25')
ws['A25'] = "3.1. Kiến thức sinh viên thu thập được:"
ws['A25'].font = font_bold

ws.merge_cells('A26:K33')
ws['A26'] = (
    "- Quy trình xây dựng phần mềm quản lý chuỗi cung ứng và việc áp dụng trong thực tế các giai đoạn khảo sát, phân tích, thiết kế, cài đặt và kiểm thử.\n"
    "- Kiến thức về quản lý chuỗi cung ứng: quy trình mua hàng, bán hàng, quản lý kho, logistics và kế toán.\n"
    "- Phát triển tất cả các thành phần của một ứng dụng Web (fullstack): Backend API, Frontend SPA, Database.\n"
    "- Nghiên cứu, triển khai các công nghệ, tính năng mới phục vụ bài toán quản lý xuất nhập kho và chuỗi cung ứng."
)
ws['A26'].font = font_normal
ws['A26'].alignment = align_left

# 3.2 Công nghệ
ws.merge_cells('A34:K34')
ws['A34'] = "3.2. Công nghệ sinh viên thu thập được:"
ws['A34'].font = font_bold

ws.merge_cells('A35:K44')
ws['A35'] = (
    "- Kiến thức về Web Client-Server Architecture, RESTful API.\n"
    "- Kiến thức về Java, hệ sinh thái Spring Boot framework (Spring Security, Spring Data JPA).\n"
    "- Kiến thức về React.js phát triển giao diện người dùng (SPA).\n"
    "- PostgreSQL cho hệ quản trị cơ sở dữ liệu quan hệ.\n"
    "- JWT (JSON Web Token) cho xác thực và phân quyền người dùng (RBAC).\n"
    "- Docker cho containerization và triển khai ứng dụng.\n"
    "- Railway cho Cloud Deployment."
)
ws['A35'].font = font_normal
ws['A35'].alignment = align_left

# 3.3 Kỹ năng
ws.merge_cells('A45:K45')
ws['A45'] = "3.3. Kỹ năng sinh viên phát triển được:"
ws['A45'].font = font_bold

ws.merge_cells('A46:K46')
ws['A46'] = (
    "- Giao tiếp, phỏng vấn khi phân tích yêu cầu, trao đổi với khách hàng, người sử dụng.\n"
    "- Tìm kiếm tổng hợp thông tin từ nhiều nguồn.\n"
    "- Khả năng quản lý thời gian và làm việc độc lập.\n"
    "- Khả năng viết báo cáo kỹ thuật chuẩn.\n"
    "- Kĩ năng trình bày, giải quyết vấn đề.\n"
)
ws['A46'].font = font_normal
ws['A46'].alignment = align_left

# 3.4 Sản phẩm kỳ vọng
ws.merge_cells('A47:K47')
ws['A47'] = "3.4. Sản phẩm kỳ vọng:"
ws['A47'].font = font_bold

ws.merge_cells('A48:K53')
ws['A48'] = (
    "- Hệ thống phần mềm quản lý chuỗi cung ứng hoàn chỉnh, hỗ trợ các nghiệp vụ: Mua hàng, Bán hàng, Quản lý Kho, Logistics, Kế toán.\n"
    "- Giao diện sản phẩm trực quan, responsive, dễ sử dụng với phân quyền theo vai trò (Admin, Quản lý, Nhân viên Kho, Kế toán, Shipper).\n"
    "- Dashboard tổng quan doanh thu, tồn kho, công nợ.\n"
    "- Tích hợp quản lý lô hàng, hạn sử dụng, và lập đợt giao hàng."
)
ws['A48'].font = font_normal
ws['A48'].alignment = align_left

# 3.5 Vấn đề thực tiễn
ws.merge_cells('A54:K54')
ws['A54'] = "3.5. Vấn đề thực tiễn đồ án giải quyết:  "
ws['A54'].font = font_bold

ws.merge_cells('A55:K61')
ws['A55'] = (
    "- Nhu cầu quản lý chuỗi cung ứng cho các doanh nghiệp vừa và nhỏ ngày càng tăng cao, đòi hỏi phần mềm phải tích hợp đồng bộ các quy trình mua hàng, bán hàng, kho vận và kế toán.\n"
    "- Quy trình nhập xuất hàng hóa thủ công dẫn đến sai sót, thiếu kiểm soát tồn kho, khó theo dõi công nợ.\n"
    "- Cần giải pháp phần mềm giúp tối ưu hóa quy trình vận hành, cung cấp dữ liệu thời gian thực cho việc ra quyết định kinh doanh.\n"
)
ws['A55'].font = font_normal
ws['A55'].alignment = align_left

# ---- Row 62: Nội dung và kế hoạch ----
ws.merge_cells('A62:K62')
ws['A62'] = "4. Các nội dung sẽ thực hiện và kế hoạch triển khai:"
ws['A62'].font = font_bold

ws.merge_cells('A63:K63')
ws['A63'] = "Lưu ý: khối lượng yêu cầu đối với đồ án tốt nghiệp hệ cử nhân là 6(0-0-12-12), i.e. 12 tiết làm việc/tuần trong 17 tuần."
ws['A63'].font = font_small_italic

# --- Nội dung 1 ---
ws.merge_cells('A64:F64')
ws['A64'] = "Nội dung 1: Tìm hiểu tổng quan về bài toán,"
ws['A64'].font = font_bold
ws['G64'] = "từ Tuần"
ws['G64'].font = font_normal
ws['H64'] = 1
ws['H64'].font = font_bold
ws['I64'] = "đến Tuần"
ws['I64'].font = font_normal
ws['J64'] = 2
ws['J64'].font = font_bold

ws.merge_cells('A65:K65')
ws['A65'] = "Chi tiết:"
ws['A65'].font = font_bold

ws.merge_cells('A66:K68')
ws['A66'] = (
    "- Xác định mục tiêu, đối tượng người dùng, vấn đề còn tồn đọng trong quản lý chuỗi cung ứng.\n"
    "- Khảo sát, trải nghiệm một số phần mềm quản lý chuỗi cung ứng: SAP, Odoo, KiotViet...\n"
    "- Xác định tiến độ cần thiết để hoàn thành đề tài.\n"
    "- Xác định các công nghệ, tài liệu liên quan phục vụ đề tài."
)
ws['A66'].font = font_normal
ws['A66'].alignment = align_left

# --- Nội dung 2 ---
ws.merge_cells('A69:F69')
ws['A69'] = "Nội dung 2: Tìm hiểu tổng quan về công nghệ liên quan, "
ws['A69'].font = font_bold
ws['G69'] = "từ Tuần"
ws['H69'] = 3
ws['H69'].font = font_bold
ws['I69'] = "đến Tuần"
ws['J69'] = 4
ws['J69'].font = font_bold

ws.merge_cells('A70:K70')
ws['A70'] = "Chi tiết:"
ws['A70'].font = font_bold

ws.merge_cells('A71:K73')
ws['A71'] = (
    "- Java, Spring Boot, Spring Security, Spring Data JPA.\n"
    "- Javascript, React.js, Vite.\n"
    "- PostgreSQL, Docker, Railway, JWT, Postman, GitHub."
)
ws['A71'].font = font_normal
ws['A71'].alignment = align_left

# --- Nội dung 3 ---
ws.merge_cells('A74:F74')
ws['A74'] = "Nội dung 3: Phân tích thiết kế,"
ws['A74'].font = font_bold
ws['G74'] = "từ Tuần"
ws['H74'] = 5
ws['H74'].font = font_bold
ws['I74'] = "đến Tuần"
ws['J74'] = 7
ws['J74'].font = font_bold

ws.merge_cells('A75:K75')
ws['A75'] = "Chi tiết:"
ws['A75'].font = font_bold

ws.merge_cells('A76:K78')
ws['A76'] = (
    "- Phân tích chi tiết các use-case hệ thống và vai trò người dùng (Admin, Quản lý Mua hàng, Quản lý Bán hàng, Nhân viên Kho, Kế toán, Shipper).\n"
    "- Phân tích luồng nghiệp vụ của từng module: Mua hàng, Bán hàng, Kho vận, Kế toán.\n"
    "- Thiết kế kiến trúc hệ thống theo mô hình phân tầng: Frontend (React), Backend (Spring Boot), Database (PostgreSQL).\n"
    "- Xác định cấu trúc cơ sở dữ liệu và các quan hệ giữa bảng dữ liệu.\n"
    "- Thiết kế giao diện người dùng (wireframe/mockup)."
)
ws['A76'].font = font_normal
ws['A76'].alignment = align_left

# --- Nội dung 4 ---
ws.merge_cells('A79:F79')
ws['A79'] = "Nội dung 4: Xây dựng chương trình,"
ws['A79'].font = font_bold
ws['G79'] = "từ Tuần"
ws['H79'] = 8
ws['H79'].font = font_bold
ws['I79'] = "đến Tuần"
ws['J79'] = 15
ws['J79'].font = font_bold

ws.merge_cells('A80:K80')
ws['A80'] = "Chi tiết:"
ws['A80'].font = font_bold

ws.merge_cells('A81:K83')
ws['A81'] = (
    "- Xây dựng các module quản lý: Sản phẩm, Khách hàng, Nhà cung cấp, Kho hàng, Đơn mua hàng (PO), Nhập kho (GR), Đơn bán hàng (SO), Xuất kho (GI), Hóa đơn.\n"
    "- Xây dựng giao diện React responsive theo thiết kế, API RESTful trả về cho Client, xử lý giao tiếp Backend và Frontend.\n"
    "- Xây dựng hệ thống xác thực JWT và phân quyền theo vai trò (RBAC).\n"
    "- Xây dựng module Kế toán: bút toán tự động, quản lý thanh toán và công nợ.\n"
    "- Xây dựng Dashboard báo cáo tổng quan doanh thu, tồn kho.\n"
    "- Triển khai ứng dụng lên Cloud (Railway)."
)
ws['A81'].font = font_normal
ws['A81'].alignment = align_left

# --- Nội dung 5 ---
ws.merge_cells('A84:F84')
ws['A84'] = "Nội dung 5: Thử nghiệm và đánh giá,"
ws['A84'].font = font_bold
ws['G84'] = "từ Tuần"
ws['H84'] = 16
ws['H84'].font = font_bold
ws['I84'] = "đến Tuần"
ws['J84'] = 17
ws['J84'].font = font_bold

ws.merge_cells('A85:K85')
ws['A85'] = "Chi tiết:"
ws['A85'].font = font_bold

ws.merge_cells('A86:K88')
ws['A86'] = (
    "- Kiểm thử đơn vị các API, kiểm thử giao diện.\n"
    "- Kiểm thử tích hợp toàn bộ luồng nghiệp vụ: PO → GR → SO → GI → Invoice → Payment.\n"
    "- Kiểm thử bảo mật API (thực thi RBAC).\n"
    "- Ghi nhận các phản hồi từ người dùng, chỉnh sửa và hoàn thiện sản phẩm."
)
ws['A86'].font = font_normal
ws['A86'].alignment = align_left

# ---- Row 89-99: Cam đoan sinh viên ----
ws.merge_cells('A89:K89')
ws['A89'] = "5. Lời cam đoan của sinh viên đã nhận được nhiệm vụ"
ws['A89'].font = font_bold

ws.merge_cells('A90:K90')
ws['A90'] = "Em xin cam kết sẽ hoàn thành các nhiệm vụ theo đúng kế hoạch."
ws['A90'].font = font_normal

ws.merge_cells('H91:K91')
ws['H91'] = "Hà Nội, ngày        tháng        năm  "
ws['H91'].font = font_italic
ws['H91'].alignment = Alignment(horizontal='center')

ws.merge_cells('H92:K92')
ws['H92'] = "Sinh viên"
ws['H92'].font = font_bold
ws['H92'].alignment = Alignment(horizontal='center')

ws.merge_cells('H93:K93')
ws['H93'] = "(Ký và ghi rõ họ tên)"
ws['H93'].font = font_italic
ws['H93'].alignment = Alignment(horizontal='center')

# Blank rows for signature
ws.merge_cells('H99:K99')
ws['H99'] = "Dương Phương Thảo"
ws['H99'].font = font_bold
ws['H99'].alignment = Alignment(horizontal='center')

# ---- Row 101-110: Xác nhận GVHD ----
ws.merge_cells('A101:K101')
ws['A101'] = "6. Xác nhận của giáo viên hướng dẫn về việc giao nhiệm vụ cho sinh viên"
ws['A101'].font = font_bold

ws.merge_cells('H102:K102')
ws['H102'] = "Hà Nội, ngày        tháng        năm  "
ws['H102'].font = font_italic
ws['H102'].alignment = Alignment(horizontal='center')

ws.merge_cells('H103:K103')
ws['H103'] = "Giảng viên hướng dẫn"
ws['H103'].font = font_bold
ws['H103'].alignment = Alignment(horizontal='center')

ws.merge_cells('H104:K104')
ws['H104'] = "(Ký và ghi rõ họ tên)"
ws['H104'].font = font_italic
ws['H104'].alignment = Alignment(horizontal='center')

ws.merge_cells('H110:K110')
ws['H110'] = "Phạm Quang Dũng"
ws['H110'].font = font_bold
ws['H110'].alignment = Alignment(horizontal='center')

# ============================================================
# Sheet 2: Bachelor (translation)
# ============================================================
ws2 = wb.create_sheet(title="Bachelor")

for col, width in col_widths.items():
    ws2.column_dimensions[col].width = width

for row in range(1, 100):
    for col in range(1, 12):
        ws2.cell(row=row, column=col).font = font_normal
        ws2.cell(row=row, column=col).alignment = align_left

# Row 1
ws2.merge_cells('J1:K1')
ws2['J1'] = "Mẫu ĐATN 02"
ws2['J1'].font = font_small_italic
ws2['J1'].alignment = Alignment(horizontal='right')

# Row 2-3
ws2.merge_cells('A2:K2')
ws2['A2'] = "HANOI UNIVERSITY OF SCIENCE AND TECHNOLOGY"
ws2['A2'].font = font_bold
ws2['A2'].alignment = align_center

ws2.merge_cells('A3:K3')
ws2['A3'] = "SCHOOL OF INFORMATION AND COMMUNICATIONS TECHNOLOGY"
ws2['A3'].font = font_bold
ws2['A3'].alignment = align_center

# Row 5
ws2.merge_cells('A5:K5')
ws2['A5'] = "THESIS ASSIGNMENT SHEET OF BACHELOR PROGRAM"
ws2['A5'].font = font_title
ws2['A5'].alignment = align_center

ws2['E6'] = "Semester"
ws2['E6'].font = font_bold
ws2['F6'] = "2025.2"
ws2['F6'].font = font_bold

# Student info
ws2.merge_cells('A8:K8')
ws2['A8'] = "Student's information"
ws2['A8'].font = font_header

ws2.merge_cells('A9:B9')
ws2['A9'] = "Full name:"
ws2['A9'].font = font_bold
ws2.merge_cells('C9:G9')
ws2['C9'] = "Duong Phuong Thao"
ws2['H9'] = "Student ID:"
ws2['H9'].font = font_bold
ws2.merge_cells('I9:K9')
ws2['I9'] = "20226001"

ws2.merge_cells('A10:B10')
ws2['A10'] = "Phone number:"
ws2['A10'].font = font_bold
ws2.merge_cells('C10:G10')
ws2['C10'] = "0913627734"
ws2['H10'] = "Class:"
ws2['H10'].font = font_bold
ws2.merge_cells('I10:K10')
ws2['I10'] = "ITE7-02-K67"

ws2.merge_cells('A11:B11')
ws2['A11'] = "Email:"
ws2['A11'].font = font_bold
ws2.merge_cells('C11:G11')
ws2['C11'] = "thao.dp226001@sis.hust.edu.vn"
ws2['H11'] = "Class code:"
ws2['H11'].font = font_bold

# Supervisor info
ws2.merge_cells('A12:K12')
ws2['A12'] = "Supervisor's information"
ws2['A12'].font = font_header

ws2.merge_cells('A13:B13')
ws2['A13'] = "Full name:"
ws2['A13'].font = font_bold
ws2.merge_cells('C13:K13')
ws2['C13'] = "Pham Quang Dung"

ws2.merge_cells('A14:B14')
ws2['A14'] = "Thesis carried out at:"
ws2['A14'].font = font_bold
ws2.merge_cells('C14:K14')
ws2['C14'] = "School of Information and Communications Technology"

ws2.merge_cells('A15:B15')
ws2['A15'] = "Time duration for thesis: "
ws2['A15'].font = font_bold
ws2['C15'] = "From "
ws2['D15'] = "24/2/2026"
ws2['H15'] = "to "
ws2['I15'] = "15/6/2026"

# Title
ws2.merge_cells('A16:K16')
ws2['A16'] = "1. Thesis's title:"
ws2['A16'].font = font_bold

ws2.merge_cells('A17:K17')
ws2['A17'] = "Smart Supply Chain Management Software"

# Field
ws2.merge_cells('A18:K18')
ws2['A18'] = "2. Thesis's field of interest:"
ws2['A18'].font = font_bold

ws2['A19'] = "-"
ws2['B19'] = "Option 1:"
ws2['B19'].font = font_bold
ws2.merge_cells('C19:K19')
ws2['C19'] = "eCommerce and Logistic"

ws2['A20'] = "-"
ws2['B20'] = "Option 2:"
ws2['B20'].font = font_bold
ws2.merge_cells('C20:K20')
ws2['C20'] = "Enterprise Software"

ws2['A21'] = "-"
ws2['B21'] = "Option 3:"
ws2['B21'].font = font_bold

ws2['A22'] = "-"
ws2.merge_cells('B22:K22')
ws2['B22'] = "If the field of interest is not listed in the available options, the supervisor may recommend a suitable topic:"
ws2['B22'].font = font_italic

# Objectives
ws2.merge_cells('A24:K24')
ws2['A24'] = "3. Objectives"
ws2['A24'].font = font_bold

ws2.merge_cells('A25:K25')
ws2['A25'] = "3.1. Knowledge acquired:"
ws2['A25'].font = font_bold

ws2.merge_cells('A26:K28')
ws2['A26'] = (
    "- Software development process for supply chain management systems: surveying, analysis, design, implementation and testing.\n"
    "- Supply chain management concepts: procurement, sales, inventory, logistics and accounting.\n"
    "- Full-stack web application development: Backend API, Frontend SPA, Database."
)
ws2['A26'].alignment = align_left

ws2.merge_cells('A29:K29')
ws2['A29'] = "3.2. Technologies acquired:"
ws2['A29'].font = font_bold

ws2.merge_cells('A30:K32')
ws2['A30'] = (
    "- Java, Spring Boot framework ecosystem (Spring Security, Spring Data JPA).\n"
    "- React.js for building user interfaces (SPA).\n"
    "- PostgreSQL, JWT, Docker, Railway Cloud Deployment."
)
ws2['A30'].alignment = align_left

ws2.merge_cells('A33:K33')
ws2['A33'] = "3.3. Skills developed:"
ws2['A33'].font = font_bold

ws2.merge_cells('A34:K34')
ws2['A34'] = (
    "- Communication and requirements analysis skills.\n"
    "- Time management and independent working.\n"
    "- Technical report writing and presentation skills."
)
ws2['A34'].alignment = align_left

ws2.merge_cells('A35:K35')
ws2['A35'] = "3.4. Products expected:"
ws2['A35'].font = font_bold

ws2.merge_cells('A36:K38')
ws2['A36'] = (
    "- A complete supply chain management system supporting: Procurement, Sales, Inventory, Logistics, Accounting.\n"
    "- Responsive web interface with role-based access control (RBAC).\n"
    "- Dashboard for revenue, inventory and debt overview."
)
ws2['A36'].alignment = align_left

ws2.merge_cells('A39:K39')
ws2['A39'] = "3.5. Practical issues solved:  "
ws2['A39'].font = font_bold

ws2.merge_cells('A40:K42')
ws2['A40'] = (
    "- Growing demand for integrated supply chain management for SMEs.\n"
    "- Manual inventory processes leading to errors and lack of control.\n"
    "- Need for real-time data solutions for business decision-making."
)
ws2['A40'].alignment = align_left

# Work packages
ws2.merge_cells('A43:K43')
ws2['A43'] = "4. Tentative work packages and plan:"
ws2['A43'].font = font_bold

ws2.merge_cells('A44:K44')
ws2['A44'] = "Note: Required credits for a bachelor-graduation thesis are 6(0-0-12-12), i.e. 12 hours per week for 17 weeks."
ws2['A44'].font = font_small_italic

# WP1
ws2.merge_cells('A45:F45')
ws2['A45'] = "Work package 1:  Literature survey,"
ws2['A45'].font = font_bold
ws2['G45'] = "from Week"
ws2['I45'] = "to Week"
ws2['H45'] = 1
ws2['H45'].font = font_bold
ws2['J45'] = 2
ws2['J45'].font = font_bold

ws2['A46'] = "Details:"
ws2['A46'].font = font_bold
ws2.merge_cells('A47:K49')
ws2['A47'] = (
    "- Define objectives, target users, and existing problems in supply chain management.\n"
    "- Survey existing SCM software: SAP, Odoo, KiotViet.\n"
    "- Identify related technologies and documentation."
)
ws2['A47'].alignment = align_left

# WP2
ws2.merge_cells('A50:F50')
ws2['A50'] = "Work package 2: Study/Research on related technologies, "
ws2['A50'].font = font_bold
ws2['G50'] = "from Week"
ws2['I50'] = "to Week"
ws2['H50'] = 3
ws2['H50'].font = font_bold
ws2['J50'] = 4
ws2['J50'].font = font_bold

ws2['A51'] = "Details:"
ws2['A51'].font = font_bold
ws2.merge_cells('A52:K54')
ws2['A52'] = "- Java, Spring Boot, Spring Security, Spring Data JPA, React.js, Vite, PostgreSQL, Docker, Railway, JWT, Postman, GitHub."
ws2['A52'].alignment = align_left

# WP3
ws2.merge_cells('A55:F55')
ws2['A55'] = "Work package 3: Analysis and design,"
ws2['A55'].font = font_bold
ws2['G55'] = "from Week"
ws2['I55'] = "to Week"
ws2['H55'] = 5
ws2['H55'].font = font_bold
ws2['J55'] = 7
ws2['J55'].font = font_bold

ws2['A56'] = "Details:"
ws2['A56'].font = font_bold
ws2.merge_cells('A57:K59')
ws2['A57'] = (
    "- Detailed use-case analysis and user role definition.\n"
    "- Business flow analysis for each module.\n"
    "- System architecture design, database schema design, UI wireframes."
)
ws2['A57'].alignment = align_left

# WP4
ws2.merge_cells('A60:F60')
ws2['A60'] = "Work package 4: Program development,"
ws2['A60'].font = font_bold
ws2['G60'] = "from Week"
ws2['I60'] = "to Week"
ws2['H60'] = 8
ws2['H60'].font = font_bold
ws2['J60'] = 15
ws2['J60'].font = font_bold

ws2['A61'] = "Details:"
ws2['A61'].font = font_bold
ws2.merge_cells('A62:K64')
ws2['A62'] = (
    "- Develop modules: Product, Customer, Supplier, Warehouse, PO, GR, SO, GI, Invoice.\n"
    "- Build responsive React UI, RESTful APIs, JWT auth with RBAC.\n"
    "- Develop Accounting module and Dashboard. Deploy to Cloud (Railway)."
)
ws2['A62'].alignment = align_left

# WP5
ws2.merge_cells('A65:F65')
ws2['A65'] = "Work package 5: Test and evaluation,"
ws2['A65'].font = font_bold
ws2['G65'] = "from Week"
ws2['I65'] = "to Week"
ws2['H65'] = 16
ws2['H65'].font = font_bold
ws2['J65'] = 17
ws2['J65'].font = font_bold

ws2['A66'] = "Details:"
ws2['A66'].font = font_bold
ws2.merge_cells('A67:K69')
ws2['A67'] = (
    "- Unit testing APIs, UI testing.\n"
    "- Integration testing: PO → GR → SO → GI → Invoice → Payment.\n"
    "- Security testing (RBAC enforcement). Collect user feedback and finalize."
)
ws2['A67'].alignment = align_left

# Commitment
ws2.merge_cells('A70:K70')
ws2['A70'] = "5. Commitment of the student who has received the assignment"
ws2['A70'].font = font_bold

ws2.merge_cells('A71:K71')
ws2['A71'] = "I commit to complete the assignment as planned above."

ws2.merge_cells('H72:K72')
ws2['H72'] = "Hanoi, Date  "
ws2['H72'].font = font_italic
ws2['H72'].alignment = Alignment(horizontal='center')

ws2.merge_cells('H73:K73')
ws2['H73'] = "Student"
ws2['H73'].font = font_bold
ws2['H73'].alignment = Alignment(horizontal='center')

ws2.merge_cells('H74:K74')
ws2['H74'] = "(Signature and full name)"
ws2['H74'].font = font_italic
ws2['H74'].alignment = Alignment(horizontal='center')

ws2.merge_cells('H80:K80')
ws2['H80'] = "Duong Phuong Thao"
ws2['H80'].font = font_bold
ws2['H80'].alignment = Alignment(horizontal='center')

# Supervisor confirmation
ws2.merge_cells('A82:K82')
ws2['A82'] = "6. Supervisor's confirmation of the student assignment"
ws2['A82'].font = font_bold

ws2.merge_cells('H83:K83')
ws2['H83'] = "Hanoi, Date  "
ws2['H83'].font = font_italic
ws2['H83'].alignment = Alignment(horizontal='center')

ws2.merge_cells('H84:K84')
ws2['H84'] = "Supervisor"
ws2['H84'].font = font_bold
ws2['H84'].alignment = Alignment(horizontal='center')

ws2.merge_cells('H85:K85')
ws2['H85'] = "(Signature and full name)"
ws2['H85'].font = font_italic
ws2['H85'].alignment = Alignment(horizontal='center')

ws2.merge_cells('H91:K91')
ws2['H91'] = "Pham Quang Dung"
ws2['H91'].font = font_bold
ws2['H91'].alignment = Alignment(horizontal='center')

# ============================================================
# Save
# ============================================================
output_path = "pgnv_datn_duongphuongthao_20226001.xlsx"
wb.save(output_path)
print(f"✅ File đã được tạo thành công: {output_path}")
